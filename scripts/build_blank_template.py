"""从根目录样例表生成空白模板，仅手动维护时运行。

python3 scripts/build_blank_template.py
"""

from __future__ import annotations

import sys
from copy import copy
from io import BytesIO
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont
from openpyxl.styles import Font

ROOT = Path(__file__).resolve().parent.parent
SOURCE = ROOT / "乌达君正7.1.xlsx"
TARGET = ROOT / "agent" / "template" / "wuda-junzheng.xlsx"

DATA_START_ROW = 3
BLANK_ROW_REF = 19
DATA_FONT = Font(name="宋体", size=10)
HEADER_INLINE = InlineFont(rFont="宋体", sz=10, b=True)
HEADER_REQUIRED_INLINE = InlineFont(rFont="宋体", sz=10, b=True, color="FFFF0000")
HEADER_FONT = Font(name="宋体", size=10, bold=True)

HEADERS = [
    (1, "提货工厂", "plain"),
    (2, "车牌号", "required"),
    (3, "挂车号", "optional"),
    (4, "司机姓名", "required"),
    (5, "身份证号", "required"),
    (6, "随车电话", "required"),
    (7, "预提数量", "required"),
    (8, "预提数量", "required"),
    (9, "流向", "required"),
]


def header_value(label: str, kind: str):
    if kind == "required":
        return CellRichText(
            TextBlock(HEADER_INLINE, label),
            TextBlock(HEADER_REQUIRED_INLINE, "（必填）"),
        )
    if kind == "optional":
        return CellRichText(
            TextBlock(HEADER_INLINE, label),
            TextBlock(HEADER_INLINE, "（选填）"),
        )
    return label


def main() -> None:
    if not SOURCE.is_file():
        print(f"找不到样例文件: {SOURCE}", file=sys.stderr)
        sys.exit(1)

    workbook = load_workbook(SOURCE, rich_text=True)
    worksheet = workbook.active
    worksheet.title = "君正派车模板"

    for col, label, kind in HEADERS:
        cell = worksheet.cell(row=2, column=col)
        cell.value = header_value(label, kind)
        if kind == "plain":
            cell.font = copy(HEADER_FONT)

    blank = {col: worksheet.cell(row=BLANK_ROW_REF, column=col).value for col in range(1, 10)}
    for row in range(DATA_START_ROW, worksheet.max_row + 1):
        for col, value in blank.items():
            cell = worksheet.cell(row=row, column=col)
            cell.value = value
            cell.font = copy(DATA_FONT)

    TARGET.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(TARGET)
    print(f"已生成空白模板: {TARGET}")


if __name__ == "__main__":
    main()
