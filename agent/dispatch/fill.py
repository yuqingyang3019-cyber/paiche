from __future__ import annotations

from copy import copy
from datetime import date, datetime
from io import BytesIO
from typing import Any
from zoneinfo import ZoneInfo

from openpyxl import load_workbook
from openpyxl.styles import Font

from .config import (
    COL_DESTINATION,
    COL_FACTORY,
    COL_ID_CARD,
    COL_NAME,
    COL_PHONE,
    COL_PLATE,
    COL_QTY_A,
    COL_QTY_B,
    DATA_START_ROW,
    DEFAULT_DESTINATION,
    DEFAULT_QUANTITY,
    FACTORY,
    TEMPLATE_PATH,
)

DATA_FONT = Font(name="宋体", size=10)


def today_cn() -> date:
    return datetime.now(ZoneInfo("Asia/Shanghai")).date()


def date_label_cn(day: date | None = None) -> str:
    current = day or today_cn()
    return f"{current.month}月{current.day}日"


def output_filename(day: date | None = None) -> str:
    current = day or today_cn()
    return f"乌达君正{current.month}.{current.day}.xlsx"


def sheet_title(day: date | None = None) -> str:
    current = day or today_cn()
    return f"君正派车模板（{current.year}年{current.month}月{current.day}日）"


def _write_cell(worksheet, row: int, col: int, value: Any) -> None:
    cell = worksheet.cell(row=row, column=col)
    cell.value = value
    cell.font = copy(DATA_FONT)


def fill_dispatch_workbook(vehicles: list[dict[str, Any]], day: date | None = None) -> tuple[bytes, str]:
    if not vehicles:
        raise ValueError("至少填写一辆车")
    if not TEMPLATE_PATH.is_file():
        raise FileNotFoundError(f"空白模板不存在: {TEMPLATE_PATH}")

    current_day = day or today_cn()
    template_bytes = TEMPLATE_PATH.read_bytes()
    workbook = load_workbook(BytesIO(template_bytes), rich_text=True)
    worksheet = workbook.active
    worksheet.title = sheet_title(current_day)

    write_row = DATA_START_ROW
    for vehicle in vehicles:
        _write_cell(worksheet, write_row, COL_FACTORY, FACTORY)
        _write_cell(worksheet, write_row, COL_PLATE, vehicle["plate"])
        _write_cell(worksheet, write_row, COL_NAME, vehicle["name"])
        _write_cell(worksheet, write_row, COL_ID_CARD, vehicle["idCard"])
        _write_cell(worksheet, write_row, COL_PHONE, int(vehicle["phone"]))
        _write_cell(worksheet, write_row, COL_QTY_A, DEFAULT_QUANTITY)
        _write_cell(worksheet, write_row, COL_QTY_B, DEFAULT_QUANTITY)
        _write_cell(worksheet, write_row, COL_DESTINATION, DEFAULT_DESTINATION)
        write_row += 1

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue(), output_filename(current_day)
