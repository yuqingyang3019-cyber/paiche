from __future__ import annotations

from datetime import date
from io import BytesIO

from unittest.mock import MagicMock, patch

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook

from dispatch.fill import fill_dispatch_workbook, output_filename
from dispatch.parse import parse_dispatch_text
from main import app

SAMPLE_TEXT = """车号：蒙L93723
姓名：康有光
电话：15164810755
身份证号：152827197608242172
车皮：15吨
自卸/国六"""

SAMPLE_VEHICLE = {
    "plate": "蒙L93723",
    "name": "康有光",
    "phone": "15164810755",
    "idCard": "152827197608242172",
}


def test_parse_single_vehicle() -> None:
    result = parse_dispatch_text(SAMPLE_TEXT)
    assert result["vehicles"] == [SAMPLE_VEHICLE]
    assert result["warnings"] == []


def test_parse_multiple_vehicles() -> None:
    text = SAMPLE_TEXT + "\n\n" + SAMPLE_TEXT.replace("蒙L93723", "蒙L61019").replace("康有光", "王建武")
    result = parse_dispatch_text(text)
    assert len(result["vehicles"]) == 2
    assert result["vehicles"][1]["plate"] == "蒙L61019"


def test_parse_missing_field() -> None:
    result = parse_dispatch_text("车号：蒙L93723\n姓名：康有光")
    assert result["vehicles"] == []
    assert result["warnings"]


def test_output_filename() -> None:
    assert output_filename(date(2026, 7, 2)) == "乌达君正7.2.xlsx"
    assert output_filename(date(2026, 12, 9)) == "乌达君正12.9.xlsx"


def test_date_label_cn() -> None:
    from dispatch.fill import date_label_cn

    assert date_label_cn(date(2026, 7, 3)) == "7月3日"


def test_fill_dispatch_workbook() -> None:
    content, filename = fill_dispatch_workbook([SAMPLE_VEHICLE], day=date(2026, 7, 2))
    assert filename == "乌达君正7.2.xlsx"

    workbook = load_workbook(BytesIO(content), rich_text=True)
    worksheet = workbook.active
    assert worksheet.title == "君正派车模板（2026年7月2日）"
    assert worksheet.cell(1, 1).value
    assert "君正报号及装车注意事项" in str(worksheet.cell(1, 1).value)
    assert worksheet.cell(3, 1).value == "乌达君正"
    assert worksheet.cell(3, 2).value == "蒙L93723"
    assert worksheet.cell(3, 4).value == "康有光"
    assert worksheet.cell(3, 5).value == "152827197608242172"
    assert worksheet.cell(3, 6).value == 15164810755
    assert worksheet.cell(3, 7).value == 37
    assert worksheet.cell(3, 9).value == "后旗团羊"
    assert worksheet.cell(3, 2).font.name == "宋体"
    assert worksheet.cell(3, 2).font.size == 10
    assert worksheet.cell(4, 2).value is None


def test_fill_requires_vehicle() -> None:
    with pytest.raises(ValueError, match="至少填写一辆车"):
        fill_dispatch_workbook([])


@pytest.fixture
def client() -> TestClient:
    return TestClient(app)


def test_health(client: TestClient) -> None:
    response = client.get("/health")
    assert response.status_code == 200
    assert response.json() == {"ok": "true"}


def test_api_parse(client: TestClient) -> None:
    response = client.post("/api/dispatch/parse", json={"text": SAMPLE_TEXT})
    assert response.status_code == 200
    payload = response.json()
    assert payload["ok"] is True
    assert payload["vehicles"][0]["plate"] == "蒙L93723"


def test_api_generate(client: TestClient) -> None:
    response = client.post("/api/dispatch/generate", json={"vehicles": [SAMPLE_VEHICLE]})
    assert response.status_code == 200
    assert "spreadsheetml.sheet" in response.headers["content-type"]
    assert len(response.content) > 1000


def test_health_with_base_path(monkeypatch) -> None:
    monkeypatch.setenv("BASE_PATH", "/paiche")
    monkeypatch.setenv("LUCHE_SKIP_ENV_LOCAL", "1")
    from main import build_app

    client = TestClient(build_app())
    response = client.get("/paiche/health")
    assert response.status_code == 200
    assert response.json() == {"ok": "true"}


MESSY_TEXT = """康有光 蒙L93723
15164810755
身份证152827197608242172
15吨自卸"""


@patch("dispatch.llm_parse._client")
def test_llm_parse_messy_text(mock_client_factory: MagicMock) -> None:
    mock_response = MagicMock()
    mock_response.choices = [
        MagicMock(
            message=MagicMock(
                content='{"vehicles":[{"plate":"蒙L93723","name":"康有光","phone":"15164810755","idCard":"152827197608242172"}],"warnings":[]}'
            )
        )
    ]
    mock_client_factory.return_value.chat.completions.create.return_value = mock_response

    with patch.dict("os.environ", {"DASHSCOPE_API_KEY": "test-key"}, clear=False):
        from dispatch.parse import parse_dispatch_text

        result = parse_dispatch_text(MESSY_TEXT)

    assert result["vehicles"] == [SAMPLE_VEHICLE]
    assert result["warnings"] == []


@patch("dispatch.llm_parse._client")
def test_llm_invalid_key_falls_back_to_regex(mock_client_factory: MagicMock) -> None:
    mock_client_factory.return_value.chat.completions.create.side_effect = RuntimeError(
        "Error code: 401 - {'error': {'code': 'invalid_api_key'}}"
    )

    with patch.dict("os.environ", {"DASHSCOPE_API_KEY": "bad-key"}, clear=False):
        from dispatch.parse import parse_dispatch_text

        result = parse_dispatch_text(SAMPLE_TEXT)

    assert result["vehicles"] == [SAMPLE_VEHICLE]
    assert result["warnings"] == []

